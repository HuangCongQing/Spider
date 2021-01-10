'''
Description: 大学相关信息爬取
Author: HCQ
Company(School): UCAS
Email: 1756260160@qq.com
Date: 2021-01-10 17:06:09
LastEditTime: 2021-01-11 01:26:45
FilePath: /Spider/practice/12阳光高考信息爬取/01get_school_info.py
'''
import requests
from bs4 import BeautifulSoup
from lxml import etree
import xlwt	# 存储
#导入线程池模块对应的类
from multiprocessing.dummy import Pool
from openpyxl import Workbook

def get_links():
    print("正在获取大学link......")
    all_urls = []
    all_names = []
    for page in range(0,141):
        #UA伪装：将对应的User-Agent封装到一个字典中
        headers = {
            'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36'
        }
        # url = 'https://gaokao.chsi.com.cn/sch/search--ss-on,option-qg,searchType-1,start-0.dhtml'
        url = 'https://gaokao.chsi.com.cn/sch/search--ss-on,searchType-1,option-qg,start-' + str(page) + '.dhtml'
        #对指定的url发起的请求对应的url是携带参数的，并且请求过程中处理了参数
        response = requests.get(url=url,headers=headers)
        response.encoding = 'utf-8' 
        page_text = response.text
        # print(page_text)
        tree = etree.HTML(page_text)
        tr_list = tree.xpath('//div[@class="yxk-table"]/table/tr')[1:]  #
        print('页面', page,  '已访问：' , len(tr_list), "大学已获得链接")
        # print(tr_list)
        urls = []
        names = []
        for tr in tr_list:
            if(tr.xpath('./td[@class="js-yxk-yxmc"]/a/@href')): # 部分学校没有链接
                name = tr.xpath('./td[@class="js-yxk-yxmc"]/a/text()')[0].replace('\n', '').replace('\r', '')
                url = 'https://gaokao.chsi.com.cn/'+tr.xpath('./td/a/@href')[0]
                urls.append(url)
                names.append(url)
                # print(name, url)
        all_urls.extend(urls)
        all_names.extend(names)
    return all_names, all_urls


def get_tabs(url):
    print("正在获取大学各版块link......")
    headers = {
        'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36'
    }
    #对指定的url发起的请求对应的url是携带参数的，并且请求过程中处理了参数
    response = requests.get(url=url,headers=headers)
    response.encoding = 'utf-8' 
    page_text = response.text
    # print(page_text)
    tree = etree.HTML(page_text)
    name = tree.xpath('//h1[@class="zx-yx-title"]/a/text()')[0]
    print(name)
    list1 = tree.xpath('//div[@class="nav-wrapper"]/div/a')
    list2 = tree.xpath('//ul[@class="nav-more-list"]/li')
    # print(list2)
    a_lists = []
    li_lists = []
    for a in list1:
        href = a.xpath('./@href')
        a_lists.append(href)
    print(len(a_lists))
    # print(a_lists[1])
    for li in list2:
        href = li.xpath('./a/@href')
        li_lists.append(href)
    print(len(li_lists))
    #               学校简介、录取规则、奖学金设置、食宿条件、联系办法、收费项目、毕业生就业、体检要求、其他
    all_list = a_lists[1] + a_lists[5] + a_lists[6] + a_lists[7] + li_lists[0]+ li_lists[2]+ li_lists[3]+ li_lists[6]+ li_lists[-1]
    # print(all_list)
    return   all_list
        

def get_contents(all_list):
    print("正在获取大学详情信息link......")
    # intro = []  # 学校简介
    # rule = [] # 录取规则
    # award = [] # 奖学金设置
    # eat = []  # 食宿条件
    # contact = [] # 联系方法
    # project = [] # 收费项目
    # employment = [] # 毕业生就业
    # physical = [] # 体检要求
    # others = [] # 其他
    containers = []
    for i in range(len(all_list)):
        print(i, '/ ' , len(all_list) ,"大学详情信息link......")
        headers = {
            'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36'
        }
        detail_url = 'https://gaokao.chsi.com.cn'+all_list[i]
        #对详情页发起请求，解析出章节内容
        detail_page_text = requests.get(url=detail_url,headers=headers).text
        #解析出详情页中相关的章节内容
        detail_soup = BeautifulSoup(detail_page_text,'lxml')
        container = detail_soup.find('div',class_='container')
        containers.append(container)
        # print(containers)
    return containers

def write_excel(schools, containers):
    print('正在保存文件.......')
    wb = Workbook()
    sheet0 = wb.create_sheet('学校统计数据', index=0) 
    head = ['学校', '学校简介', '录取规则', '奖学金设置', '食宿条件', '联系方法', '收费项目', '毕业生就业', '体检要求', '其他']
    col0 = schools
    print('学校总数：', len(col0))
    for i, item in enumerate(head):
        sheet0.cell(row = 1,column=i+1,value=item) # 表头
    for i in range(len(col0)): #2820
        sheet0.cell(row = i+2,column=1,value=col0[i])
        sheet0.cell(row = i+2,column=2,value=str(containers[i][0]))
        sheet0.cell(row = i+2,column=3,value=str(containers[i][1]))
        sheet0.cell(row = i+2,column=4,value=str(containers[i][2]))
        sheet0.cell(row = i+2,column=5,value=str(containers[i][3]))
        sheet0.cell(row = i+2,column=6,value=str(containers[i][4]))
        sheet0.cell(row = i+2,column=7,value=str(containers[i][5]))
        sheet0.cell(row = i+2,column=8,value=str(containers[i][6]))
        sheet0.cell(row = i+2,column=9,value=str(containers[i][7]))
        sheet0.cell(row = i+2,column=10,value=str(containers[i][8]))
        #保存文件
    wb.save('学校统计数据.xlsx') 
    # #创建工作簿
    # workbook = xlwt.Workbook(encoding='utf-8')  
    # #创建sheet
    # data_sheet = workbook.add_sheet('学校统计数据')
    # # 测试数据
    # col0 = schools
    # for i in range(len(col0)):
    #     data_sheet.write(i, 0,col0[i])  # set_style 设置属性
    #     data_sheet.write(i, 1,  str(containers[i][0]))
    #     data_sheet.write(i, 2,  str(containers[i][1]))
    #     data_sheet.write(i, 3, str(containers[i][2]))
    #     data_sheet.write(i, 4, str(containers[i][3]))
    #     data_sheet.write(i, 5, str(containers[i][4]))
    #     data_sheet.write(i, 6, str(containers[i][5]))
    #     data_sheet.write(i, 7, str(containers[i][6]))
    #     data_sheet.write(i, 8, str(containers[i][7]))
    #     data_sheet.write(i, 9, str(containers[i][8]))
    # #保存文件
    # workbook.save('大学详情数据.xls')	

if __name__ == "__main__":
    schools, urls = get_links()
    # print(schools, urls)
    print("大学总数：", len(schools))
    # urls = ['https://gaokao.chsi.com.cn/sch/schoolInfoMain--schId-1.dhtml', 'https://gaokao.chsi.com.cn/sch/schoolInfoMain--schId-1.dhtml']
    # schools = ['北大', '清华']
    # all_list = get_tabs(url)

    #实例化一个线程池对象
    pool = Pool(20) # 4 个进程
    #将列表中每一个列表元素传递给get_page进行处理。
    content1 = pool.map(get_tabs,urls)  # 函数 ，参数,   若有返回值一定是个列表
    # print(content1)
    pool.close()
    pool.join()

    # containers = get_contents(all_list)
    #实例化一个线程池对象
    pool2 = Pool(20) # 4 个进程
    #将列表中每一个列表元素传递给get_page进行处理。
    containers = pool2.map(get_contents,content1)  # 函数 ，参数,   若有返回值一定是个列表
    # print(containers)
    pool2.close()
    pool2.join()

    write_excel(schools, containers)