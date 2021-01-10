'''
Description: 
Author: HCQ
Company(School): UCAS
Email: 1756260160@qq.com
Date: 2021-01-05 10:33:12
LastEditTime: 2021-01-11 01:05:21
FilePath: /Spider/文件操作/excel/04openpyxl存储.py
'''
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook



def write_excel(fileName):
    wb = Workbook()
    sheet0 = wb.create_sheet('123', index=0)
    head = ['学校', '学校简介', '录取规则', '奖学金设置', '食宿条件', '联系方法', '收费项目', '毕业生就业', '体检要求', '其他']
    for i, item in enumerate(head):
        sheet0.cell(row = 1,column=i+1,value=item)
    # sheet0.cell(row = 1,column=1,value='姓名')
    # sheet0.cell(row = 1,column=2,value='年龄')
    # #写入表头
    # dilei_head = ['比对信息','申请用地面积','图斑面积','图斑地类面积','线状地物、田坎面积','其中国有土地','其中集体土地','已批面积','报批面积']
    # dilei_col = ['总面积','农用地','耕地','其中：水田','其中：水浇地','园地','林地','带K面积','其他(包含养殖)','坑塘水面','建设用地','未利用地']
    # gengdi_head = ['耕地','水田','水浇地','旱地','1等','2等','3等','4等','5等','6等','7等','8等','9等','10等','11等','12等','13等','14等','15等']
    # gengdi_col = ['占压耕地','还原耕地','线状地物、田坎','2015后新增','2015后减少','分析面积','申请面积']
    # sheet0Name = '地类分析'
    # sheet1Name = '耕地分析'
    # sheet0 = wb.create_sheet(sheet0Name.decode('gbk'), index=0)
    # for i, item in enumerate(dilei_head):
    #     sheet0.cell(row = 1,column=i+2,value=item.decode('gbk'))
    # for i, item in enumerate(dilei_col):
    #     sheet0.cell(row = i+2,column=1,value=item.decode('gbk'))
    # sheet1 = wb.create_sheet(sheet1Name.decode('gbk'), index=1)
    # for i, item in enumerate(gengdi_head):
    #     sheet1.cell(row = 1,column=i+2,value=item.decode('gbk'))
    # for i, item in enumerate(gengdi_col):
    #     sheet1.cell(row = i+2,column=1,value=item.decode('gbk'))
    # #写入数据
    # for i, item in enumerate(data0):
    #     i = i+2
    #     for j,val in enumerate(item):
    #         sheet0.cell(row = i,column=j+2,value=val)
    # for i, item in enumerate(data1):
    #     i = i+2
    #     for j,val in enumerate(item):
    #         sheet1.cell(row = i,column=j+2,value=val)
        
    wb.save(fileName + '.xlsx') 


if __name__ == "__main__":
    write_excel('04openpyxl存储test')