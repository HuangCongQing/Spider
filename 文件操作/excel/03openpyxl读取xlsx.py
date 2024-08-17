'''
Description: https://blog.csdn.net/qq_46106857/article/details/127727827?
Author: HCQ
Company(School): UCAS
Email: 1756260160@qq.com
Date: 2021-01-05 10:33:12
LastEditTime: 2024-08-18 03:01:04
FilePath: /Spider/文件操作/excel/03openpyxl读取.py
'''
import openpyxl
# #定义一个空列表
# stu_num=[]
# #打开目标execl，这里注意openpyxl能读取的execl后缀名是'.xlsx'文件
# workbook1=openpyxl.load_workbook(r'/home/hcq/python/Spider/文件操作/excel/college.xlsx')
# #选定目标sheet
# worksheet1 = workbook1.active
# for cell in worksheet1['A']:
#     # print(cell.value)
#     stu_num.append(cell.value)#这里用循环把A列每个cell的值写入开始定义的空列表
# print(stu_num) # 获取一列

# 标准开头格式
from openpyxl import load_workbook
# 打开工作薄
wb = load_workbook(r'/home/hcq/project/zhiyuan/Spider/practice/31土壤信息获取/test.xlsx')
# 获取当前所有的sheet
sheets = wb.worksheets
# 读取第一个sheet表格
sheet1 = sheets[0]


# 3、读取第一行的内容
# 一般xlsx第一列都是数据的字段，所以后续处理需要的话，还是先读取这一行再说
row_list = []
for row in sheet1[1]:
    row_list.append(row.value)

# 4、读取每一行的内容
# 获取行数
max_row_num = sheet1.max_row

for i in range(2, max_row_num + 1):
    row_list2 = []
    for row in sheet1[i]:
        row_list2.append(row.value)



# 5、读取每一列的内容
lists = []
for column in sheet1.columns:
    column_list = []
    # 将该列的每一行对应的行数据依次添加到list集合，相当于获取到这一列的数据
    for cell in column:
        column_list.append(cell.value)
    # 再将每一列存到lists列表里
    lists.append(column_list)


# 如果要看某一列的数据
# 看第1列的内容
print(lists[1])
